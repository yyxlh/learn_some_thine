#!/usr/bin/python3
# -*- coding: UTF-8 -*-
import openpyxl
from openpyxl.styles import NamedStyle,Font

#read from excel

wb=openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet.title)

print(sheet['A1'].value)
print(sheet['b1'].value)
print(sheet['c1'].value)
print(sheet['c1'].row)
print(sheet['c1'].column)
print(sheet['c1'].coordinate)

print(sheet.get_cell_collection())
print(sheet.cell(row=2,column=3).value)

for rowobj in sheet['A1':'C3']:
    for cell in rowobj:
        print(cell.value)
    print('-----END of ROW----')

# write to excel
wb=openpyxl.Workbook()
sheet=wb.get_active_sheet()
sheet.title='yy sheet'
wb.create_sheet(index=0,title='sheet1')
wb.create_sheet(index=2,title='sheet2')
print(wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('sheet1'))
print(wb.get_sheet_names())
wb.save('write_xls.xlsx')

#set the style
#使用公式
#合并单元格
#冻结行或者列
#生产图标chart
